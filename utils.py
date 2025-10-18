from __future__ import annotations
import io, json, re
from typing import Any, Dict, Optional, List
from datetime import datetime, date

import pandas as pd
from models import Project
from prompt_templates import build_system_prompt

# ----------------------------
# JSON serialization helpers
# ----------------------------
class DateTimeEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles date and datetime objects"""
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return super().default(obj)

def _serialize_for_json(obj: Any) -> Any:
    """Recursively convert dates to ISO strings for JSON serialization"""
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {k: _serialize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_serialize_for_json(item) for item in obj]
    elif hasattr(obj, 'model_dump'):
        # Pydantic model
        return _serialize_for_json(obj.model_dump())
    return obj

# ----------------------------
# Project I/O (Complete State)
# ----------------------------
def save_project(path: str, project: Project) -> None:
    """Save project to JSON file with date serialization"""
    project_dict = project.model_dump()
    serialized = _serialize_for_json(project_dict)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(serialized, f, ensure_ascii=False, indent=2, cls=DateTimeEncoder)

def save_complete_state(
    path: str,
    project: Project,
    schedule_payload: Optional[Dict[str, Any]] = None,
    generated_schedule: Optional[Dict[str, Any]] = None,
    generated_entries: Optional[List[Any]] = None,
    llm_conversation: Optional[List[Dict[str, Any]]] = None,
    schedule_manager_state: Optional[Any] = None
) -> None:
    """
    Save complete application state including:
    - Project configuration (employees, shifts, rules, LLM config)
    - Schedule manager state (uploaded/generated entries, conflicts)
    - Imported schedule data (schedule_payload)
    - Generated schedule output
    - Generated schedule entries
    - LLM conversation history
    """
    state = {
        "version": "2.0",  # Bumped version for schedule manager support
        "saved_at": datetime.now().isoformat(),
        "project": project.model_dump(),
        "schedule_payload": schedule_payload,
        "generated_schedule": generated_schedule,
        "generated_entries": [e.model_dump() if hasattr(e, 'model_dump') else e for e in (generated_entries or [])],
        "llm_conversation": llm_conversation or [],
        "schedule_manager_state": schedule_manager_state.model_dump() if schedule_manager_state and hasattr(schedule_manager_state, 'model_dump') else None
    }

    # Serialize dates
    serialized = _serialize_for_json(state)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(serialized, f, ensure_ascii=False, indent=2, cls=DateTimeEncoder)

def load_project_dict(data: Dict[str, Any]) -> Project:
    """Load project from dict, handling both simple and complete state formats"""
    # Check if this is a complete state file (new format)
    if "project" in data and "version" in data:
        return Project.model_validate(data["project"])
    # Otherwise treat as simple project file (old format)
    return Project.model_validate(data)

def load_complete_state(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Load complete application state from dict.
    Returns dict with keys: project, schedule_payload, generated_schedule, generated_entries,
    llm_conversation, schedule_manager_state
    """
    # Check if this is a complete state file
    if "version" in data and "project" in data:
        from models import ScheduleEntry, ScheduleState, GeneratedScheduleEntry

        # Convert generated_entries back to ScheduleEntry objects
        generated_entries = []
        if data.get("generated_entries"):
            for entry_data in data["generated_entries"]:
                try:
                    generated_entries.append(ScheduleEntry.model_validate(entry_data))
                except:
                    pass  # Skip invalid entries

        # Restore schedule manager state if present (v2.0+)
        schedule_manager_state = None
        if data.get("schedule_manager_state"):
            try:
                schedule_manager_state = ScheduleState.model_validate(data["schedule_manager_state"])
            except:
                pass  # Fall back to None if invalid

        return {
            "project": Project.model_validate(data["project"]),
            "schedule_payload": data.get("schedule_payload"),
            "generated_schedule": data.get("generated_schedule"),
            "generated_entries": generated_entries,
            "llm_conversation": data.get("llm_conversation", []),
            "schedule_manager_state": schedule_manager_state
        }
    else:
        # Old format - just project
        return {
            "project": Project.model_validate(data),
            "schedule_payload": None,
            "generated_schedule": None,
            "generated_entries": [],
            "llm_conversation": [],
            "schedule_manager_state": None
        }

# ----------------------------
# Prompt compiler
# ----------------------------
def compile_prompt(project: Project, schedule_payload: Optional[Dict[str, Any]] = None, today_iso: Optional[str] = None) -> str:
    return build_system_prompt(project, schedule_payload=schedule_payload, today_iso=today_iso)

# ----------------------------
# OpenAI-compatible call
# ----------------------------
import requests
def call_llm(
    prompt: str,
    endpoint: str,
    api_key: str,
    model: str,
    temperature: float = 0.2,
    json_mode: bool = False,
    timeout: int = 60,
) -> str:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    url = endpoint.rstrip("/") + "/v1/chat/completions"
    payload: Dict[str, Any] = {
        "model": model,
        "messages": [{"role": "system", "content": prompt}, {"role": "user", "content": "Produce the schedule now."}],
        "temperature": temperature,
    }
    if json_mode:
        payload["response_format"] = {"type": "json_object"}
    resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    return data["choices"][0]["message"]["content"]

# ----------------------------
# Schedule file parsing
# ----------------------------

DATE_COL_CANDIDATES = ["date", "datum", "day", "datum_tag", "startdatum"]
FROM_COL_CANDIDATES = ["date_from", "from", "von", "startdatum"]
TO_COL_CANDIDATES   = ["date_to", "to", "bis", "enddatum"]
EMP_COL_CANDIDATES  = ["employee", "employee_id", "name", "mitarbeiter", "person", "mitglied"]
EMAIL_COL_CANDIDATES = ["email", "e_mail", "e_mail_geschaftlich", "business_email"]
GROUP_COL_CANDIDATES = ["group", "gruppe", "team", "abteilung"]
ROLE_COL_CANDIDATES = ["role", "position", "funktion", "bereich", "bezeichnung"]
SHIFT_COL_CANDS     = ["shift", "schicht"]
START_TIME_CANDS    = ["start", "start_time", "beginn", "startzeit"]
END_TIME_CANDS      = ["end", "end_time", "ende", "endzeit"]
TYPE_COL_CANDS      = ["type", "event", "reason", "typ", "art", "grund_fur_arbeitsfreie_zeit"]
NOTES_COL_CANDS     = ["notes", "note", "bemerkung", "comment", "notizen"]
COLOR_COL_CANDS     = ["color", "colour", "themenfarbe", "farbe"]
BREAK_COL_CANDS     = ["break", "pause", "unpaid_break", "unbezahlte_pause_minuten"]
SHARED_COL_CANDS    = ["shared", "geteilt"]

def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    def normalize_col(c: str) -> str:
        # Convert to lowercase and strip
        s = c.strip().lower()
        # Replace German umlauts before removing special chars
        s = s.replace('ä', 'a').replace('ö', 'o').replace('ü', 'u').replace('ß', 'ss')
        # Replace non-alphanumeric with underscore
        s = re.sub(r"[^a-z0-9]+", "_", s)
        # Remove leading/trailing underscores
        return s.strip("_")
    m = {c: normalize_col(c) for c in df.columns}
    return df.rename(columns=m)

def _pick(df: pd.DataFrame, cands: List[str]) -> Optional[str]:
    for c in cands:
        if c in df.columns:
            return c
    return None

def _parse_date(val) -> Optional[date]:
    if pd.isna(val):
        return None
    if isinstance(val, (pd.Timestamp, )):
        return val.date()
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    try:
        return pd.to_datetime(s, dayfirst=False, utc=False).date()
    except Exception:
        try:
            return pd.to_datetime(s, dayfirst=True, utc=False).date()
        except Exception:
            return None

def _read_schedule_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    name = filename.lower()
    bio = io.BytesIO(file_bytes)
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(bio)
    else:
        # let pandas sniff separators; fallback to comma
        try:
            df = pd.read_csv(bio)
        except Exception:
            bio.seek(0)
            df = pd.read_csv(bio, sep=";")
    return _normalize_cols(df)

def _expand_ranges(df: pd.DataFrame, col_from: str, col_to: str) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        d1 = _parse_date(r.get(col_from))
        d2 = _parse_date(r.get(col_to))
        if not d1 and not d2:
            # nothing to expand, skip
            continue
        if d1 and not d2:
            d2 = d1
        if d2 and not d1:
            d1 = d2
        if d1 and d2 and d2 < d1:
            d1, d2 = d2, d1
        # n.b. guard long ranges to 366 days
        if d1 and d2:
            rng = pd.date_range(d1, d2, freq="D")[:366]
            for d in rng:
                rr = dict(r)
                rr["date"] = d.date().isoformat()
                rows.append(rr)
    if rows:
        out = pd.DataFrame(rows)
        return _normalize_cols(out)
    return df

def parse_schedule_to_payload(file_bytes: bytes, filename: str, today: date) -> Dict[str, Any]:
    """
    Returns a compact JSON payload with two top-level keys:
      - past_entries[]    (date < today)
      - future_entries[]  (date >= today)
    Each entry carries selected normalized fields.
    """
    df = _read_schedule_file(file_bytes, filename)
    # try to create a single date column
    col_date = _pick(df, DATE_COL_CANDIDATES)
    col_from = _pick(df, FROM_COL_CANDIDATES)
    col_to   = _pick(df, TO_COL_CANDIDATES)

    if not col_date and (col_from or col_to):
        df = _expand_ranges(df, col_from or col_to, col_to or col_from)
        col_date = "date"

    if not col_date and "date" not in df.columns:
        # last resort: look for something that parses as date in first non-null column
        for c in df.columns:
            maybe = _parse_date(df[c].dropna().iloc[0]) if not df[c].dropna().empty else None
            if maybe:
                col_date = c
                break

    if not col_date:
        raise ValueError("Could not detect a date/date_from/date_to column in the uploaded schedule file.")

    # pick optional fields
    col_emp  = _pick(df, EMP_COL_CANDIDATES)
    col_role = _pick(df, ROLE_COL_CANDIDATES)
    col_shift= _pick(df, SHIFT_COL_CANDS)
    col_start= _pick(df, START_TIME_CANDS)
    col_end  = _pick(df, END_TIME_CANDS)
    col_type = _pick(df, TYPE_COL_CANDS)
    col_notes= _pick(df, NOTES_COL_CANDS)

    # normalize row dicts
    def make_row(sr: pd.Series) -> Dict[str, Any]:
        d = _parse_date(sr[col_date])
        return {
            "date": d.isoformat() if d else None,
            **({"employee": str(sr[col_emp]).strip()} if col_emp else {}),
            **({"role": str(sr[col_role]).strip()} if col_role else {}),
            **({"shift": str(sr[col_shift]).strip()} if col_shift else {}),
            **({"start": str(sr[col_start]).strip()} if col_start else {}),
            **({"end": str(sr[col_end]).strip()} if col_end else {}),
            **({"type": str(sr[col_type]).strip()} if col_type else {}),
            **({"notes": str(sr[col_notes]).strip()} if col_notes else {}),
        }

    rows: List[Dict[str, Any]] = [make_row(r) for _, r in df.iterrows() if _parse_date(r[col_date]) is not None]

    # split by today
    past: List[Dict[str, Any]] = []
    future: List[Dict[str, Any]] = []
    for r in rows:
        d = pd.to_datetime(r["date"]).date()
        if d < today:
            past.append(r)
        else:
            future.append(r)

    # compact / cap to avoid massive prompts
    past_sorted = sorted(past, key=lambda x: x["date"], reverse=True)
    future_sorted = sorted(future, key=lambda x: x["date"])

    # Limit payload sizes sensibly (adjust as needed)
    past_cap = past_sorted[:800]
    future_cap = future_sorted[:800]

    # derive lightweight fairness hints (last 14 days)
    last_14 = [r for r in past_sorted if (today - pd.to_datetime(r["date"]).date()).days <= 14]
    fairness_hints = _compute_fairness_hints(last_14)

    return {
        "meta": {
            "source_filename": filename,
            "today": today.isoformat(),
            "rows_in_file": len(rows),
            "rows_past_included": len(past_cap),
            "rows_future_included": len(future_cap),
        },
        "past_entries": past_cap,
        "future_entries": future_cap,
        "fairness_hints": fairness_hints,  # e.g., recent late/pikett streaks per employee
    }

def _compute_fairness_hints(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Very lightweight heuristic summaries for rotation fairness:
    counts of entries labeled as 'late' or 'pikett' in the recent window.
    """
    counts: Dict[str, Dict[str, int]] = {}
    for r in rows:
        emp = r.get("employee") or "unknown"
        role = (r.get("role") or "").lower()
        shift = (r.get("shift") or "").lower()
        label_late = "late" in shift or "spät" in shift or ("10:00" in (r.get("start") or "") and "19:00" in (r.get("end") or ""))
        label_pikett = "pikett" in role or "pikett" in shift or "piket" in role or "piket" in shift
        if emp not in counts:
            counts[emp] = {"late_14d": 0, "pikett_14d": 0}
        if label_late:
            counts[emp]["late_14d"] += 1
        if label_pikett:
            counts[emp]["pikett_14d"] += 1
    return counts

# ----------------------------
# Dual-file Teams import
# ----------------------------
def parse_dual_schedule_files(
    shifts_file_bytes: bytes,
    shifts_filename: str,
    timeoff_file_bytes: Optional[bytes],
    timeoff_filename: Optional[str],
    today: date
) -> Dict[str, Any]:
    """
    Parse Microsoft Teams Shifts and Time-Off exports.

    Args:
        shifts_file_bytes: Shifts Excel file
        shifts_filename: Name of shifts file
        timeoff_file_bytes: Optional time-off Excel file
        timeoff_filename: Optional name of time-off file
        today: Reference date for splitting past/future

    Returns:
        Unified payload with past/future entries categorized by type
    """
    all_entries: List[Dict[str, Any]] = []

    # Parse shifts file
    df_shifts = _read_schedule_file(shifts_file_bytes, shifts_filename)
    shifts_entries = _parse_teams_file(df_shifts, entry_type="shift")
    all_entries.extend(shifts_entries)

    # Parse time-off file if provided
    if timeoff_file_bytes and timeoff_filename:
        df_timeoff = _read_schedule_file(timeoff_file_bytes, timeoff_filename)
        timeoff_entries = _parse_teams_file(df_timeoff, entry_type="time_off")
        all_entries.extend(timeoff_entries)

    # Split by today
    past: List[Dict[str, Any]] = []
    future: List[Dict[str, Any]] = []

    for entry in all_entries:
        entry_date = pd.to_datetime(entry["start_date"]).date()
        if entry_date < today:
            past.append(entry)
        else:
            future.append(entry)

    # Sort and cap
    past_sorted = sorted(past, key=lambda x: x["start_date"], reverse=True)
    future_sorted = sorted(future, key=lambda x: x["start_date"])

    past_cap = past_sorted[:800]
    future_cap = future_sorted[:800]

    # Compute fairness hints from last 14 days
    last_14 = [r for r in past_sorted if (today - pd.to_datetime(r["start_date"]).date()).days <= 14]
    fairness_hints = _compute_fairness_hints(last_14)

    return {
        "meta": {
            "source_shifts_file": shifts_filename,
            "source_timeoff_file": timeoff_filename if timeoff_filename else "Not provided",
            "today": today.isoformat(),
            "total_entries": len(all_entries),
            "rows_past_included": len(past_cap),
            "rows_future_included": len(future_cap),
        },
        "past_entries": past_cap,
        "future_entries": future_cap,
        "fairness_hints": fairness_hints,
    }

def _parse_teams_file(df: pd.DataFrame, entry_type: str = "shift") -> List[Dict[str, Any]]:
    """Parse a Teams export file (shifts or time-off) into entries"""
    # Pick columns
    col_emp = _pick(df, EMP_COL_CANDIDATES)
    col_email = _pick(df, EMAIL_COL_CANDIDATES)
    col_group = _pick(df, GROUP_COL_CANDIDATES)
    col_start_date = _pick(df, FROM_COL_CANDIDATES) or _pick(df, DATE_COL_CANDIDATES)
    col_start_time = _pick(df, START_TIME_CANDS)
    col_end_date = _pick(df, TO_COL_CANDIDATES)
    col_end_time = _pick(df, END_TIME_CANDS)
    col_color = _pick(df, COLOR_COL_CANDS)
    col_label = _pick(df, ROLE_COL_CANDIDATES)
    col_break = _pick(df, BREAK_COL_CANDS)
    col_notes = _pick(df, NOTES_COL_CANDS)
    col_shared = _pick(df, SHARED_COL_CANDS)
    col_reason = _pick(df, TYPE_COL_CANDS)

    if not col_emp or not col_start_date:
        raise ValueError(f"Could not detect employee or date columns in {entry_type} file")

    entries = []
    for _, row in df.iterrows():
        start_date = _parse_date(row[col_start_date])
        end_date = _parse_date(row[col_end_date]) if col_end_date else start_date

        if not start_date:
            continue

        # Extract employee name (remove quotes if present)
        emp_name = str(row[col_emp]).strip().strip('"')

        # Parse color code from Teams format (e.g., "1. Weiß" -> "1")
        color_code = None
        if col_color and not pd.isna(row[col_color]):
            color_str = str(row[col_color]).strip()
            if color_str and color_str[0].isdigit():
                color_code = color_str.split('.')[0]

        entry = {
            "employee": emp_name,
            "email": str(row[col_email]).strip() if col_email and not pd.isna(row[col_email]) else None,
            "group": str(row[col_group]).strip() if col_group and not pd.isna(row[col_group]) else None,
            "start_date": start_date.isoformat(),
            "start_time": str(row[col_start_time]).strip() if col_start_time and not pd.isna(row[col_start_time]) else None,
            "end_date": end_date.isoformat() if end_date else start_date.isoformat(),
            "end_time": str(row[col_end_time]).strip() if col_end_time and not pd.isna(row[col_end_time]) else None,
            "color_code": color_code,
            "label": str(row[col_label]).strip() if col_label and not pd.isna(row[col_label]) else None,
            "unpaid_break": int(row[col_break]) if col_break and not pd.isna(row[col_break]) else None,
            "notes": str(row[col_notes]).strip() if col_notes and not pd.isna(row[col_notes]) else None,
            "shared": str(row[col_shared]).strip() if col_shared and not pd.isna(row[col_shared]) else "1. Geteilt",
            "entry_type": entry_type,
            "reason": str(row[col_reason]).strip() if col_reason and not pd.isna(row[col_reason]) else None,
        }
        entries.append(entry)

    return entries

# ----------------------------
# Single-file Teams import (multi-sheet Excel)
# ----------------------------
def parse_teams_excel_multisheet(
    file_bytes: bytes,
    filename: str,
    today: date
) -> Dict[str, Any]:
    """
    Parse a single Microsoft Teams Shifts Excel export with multiple sheets.

    Expected sheets (German names, 3 out of 5 are relevant):
    - "Schichten" (Shifts) - actual shift assignments
    - "Arbeitsfreie Zeit" (Time-Off) - vacation, sick leave, etc.
    - "Mitglieder" (Members) - employee list with emails

    The other two sheets are ignored.

    Args:
        file_bytes: Multi-sheet Teams Excel file
        filename: Name of the Excel file
        today: Reference date for splitting past/future

    Returns:
        Unified payload with past/future entries, member data, and metadata
    """
    # Read all sheets from Excel file
    bio = io.BytesIO(file_bytes)
    excel_file = pd.ExcelFile(bio)

    # Detect relevant sheets (case-insensitive, flexible matching)
    shifts_sheet = None
    timeoff_sheet = None
    members_sheet = None

    for sheet_name in excel_file.sheet_names:
        normalized = sheet_name.lower().strip()
        # Prioritize exact matches, then substring matches
        if normalized == "schichten" or normalized == "shifts":
            shifts_sheet = sheet_name
        elif "arbeitsfreie" in normalized or ("time" in normalized and "off" in normalized):
            timeoff_sheet = sheet_name
        elif normalized == "mitglieder" or normalized == "members":
            members_sheet = sheet_name
        # Fallback to substring matching only if exact match not found
        elif not shifts_sheet and ("schichten" in normalized or "shifts" in normalized):
            shifts_sheet = sheet_name
        elif not members_sheet and ("mitglied" in normalized or "member" in normalized):
            members_sheet = sheet_name

    if not shifts_sheet:
        raise ValueError(f"Could not find 'Schichten' (Shifts) sheet in Excel file. Available sheets: {excel_file.sheet_names}")

    all_entries: List[Dict[str, Any]] = []
    members_data: List[Dict[str, str]] = []

    # Parse Schichten (Shifts)
    df_shifts = pd.read_excel(bio, sheet_name=shifts_sheet)
    df_shifts = _normalize_cols(df_shifts)
    shifts_entries = _parse_teams_file(df_shifts, entry_type="shift")
    all_entries.extend(shifts_entries)

    # Parse Arbeitsfreie Zeit (Time-Off) if available
    if timeoff_sheet:
        bio.seek(0)  # Reset stream
        df_timeoff = pd.read_excel(bio, sheet_name=timeoff_sheet)
        df_timeoff = _normalize_cols(df_timeoff)
        timeoff_entries = _parse_teams_file(df_timeoff, entry_type="time_off")
        all_entries.extend(timeoff_entries)

    # Parse Mitglieder (Members) if available
    if members_sheet:
        bio.seek(0)  # Reset stream
        df_members = pd.read_excel(bio, sheet_name=members_sheet)
        df_members = _normalize_cols(df_members)

        col_emp = _pick(df_members, EMP_COL_CANDIDATES)
        col_email = _pick(df_members, EMAIL_COL_CANDIDATES)

        if col_emp and col_email:
            for _, row in df_members.iterrows():
                if not pd.isna(row[col_emp]) and not pd.isna(row[col_email]):
                    members_data.append({
                        "name": str(row[col_emp]).strip().strip('"'),
                        "email": str(row[col_email]).strip()
                    })

    # Split by today
    past: List[Dict[str, Any]] = []
    future: List[Dict[str, Any]] = []

    for entry in all_entries:
        entry_date = pd.to_datetime(entry["start_date"]).date()
        if entry_date < today:
            past.append(entry)
        else:
            future.append(entry)

    # Sort and cap
    past_sorted = sorted(past, key=lambda x: x["start_date"], reverse=True)
    future_sorted = sorted(future, key=lambda x: x["start_date"])

    past_cap = past_sorted[:800]
    future_cap = future_sorted[:800]

    # Compute fairness hints from last 14 days
    last_14 = [r for r in past_sorted if (today - pd.to_datetime(r["start_date"]).date()).days <= 14]
    fairness_hints = _compute_fairness_hints(last_14)

    return {
        "meta": {
            "source_filename": filename,
            "today": today.isoformat(),
            "sheets_found": {
                "shifts": shifts_sheet,
                "timeoff": timeoff_sheet if timeoff_sheet else "Not found",
                "members": members_sheet if members_sheet else "Not found"
            },
            "total_entries": len(all_entries),
            "rows_past_included": len(past_cap),
            "rows_future_included": len(future_cap),
            "members_count": len(members_data),
        },
        "past_entries": past_cap,
        "future_entries": future_cap,
        "fairness_hints": fairness_hints,
        "members": members_data,
    }

# ----------------------------
# Employee auto-population from Excel
# ----------------------------
def find_duplicate_employee(project: Project, name: str, email: str) -> Optional[Any]:
    """
    Check if an employee with the same name AND email already exists.

    Args:
        project: Current project with employees list
        name: Employee name to check
        email: Employee email to check

    Returns:
        Existing Employee object if found, None otherwise
    """
    # Normalize for comparison
    name_normalized = name.strip().lower()
    email_normalized = email.strip().lower() if email else ""

    for emp in project.employees:
        emp_name_normalized = emp.name.strip().lower()
        emp_email_normalized = emp.email.strip().lower() if emp.email else ""

        # Match if both name AND email are the same
        if emp_name_normalized == name_normalized and emp_email_normalized == email_normalized:
            return emp

    return None

def auto_populate_employees_from_members(
    project: Project,
    members_data: List[Dict[str, str]]
) -> Dict[str, Any]:
    """
    Auto-populate employees from Excel member data with duplicate detection.

    Args:
        project: Current project (will be modified in place)
        members_data: List of member dicts with 'name' and 'email' keys

    Returns:
        Summary dict with counts of added/existing/updated employees
    """
    from models import Employee

    added_count = 0
    existing_count = 0
    added_employees = []
    existing_employees = []

    for member in members_data:
        name = member.get("name", "").strip()
        email = member.get("email", "").strip()

        if not name or not email:
            continue

        # Check for duplicates
        existing = find_duplicate_employee(project, name, email)

        if existing:
            # Employee already exists
            existing_count += 1
            existing_employees.append({
                "name": name,
                "email": email,
                "status": "already_exists"
            })
        else:
            # Create new employee
            new_emp = Employee(
                id=f"emp_{len(project.employees) + 1}_{name.lower().replace(' ', '_')}",
                name=name,
                email=email,
                roles=[],  # Will be inferred from shifts later
                percent=100,  # Default to full-time
            )
            project.employees.append(new_emp)
            added_count += 1
            added_employees.append({
                "name": name,
                "email": email,
                "status": "added",
                "id": new_emp.id
            })

    return {
        "added_count": added_count,
        "existing_count": existing_count,
        "total_members": len(members_data),
        "added_employees": added_employees,
        "existing_employees": existing_employees,
    }

def generate_schedule_preview(
    schedule_payload: Dict[str, Any],
    employee_changes: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Generate a preview of the parsed schedule and employee changes.

    Args:
        schedule_payload: The parsed schedule data
        employee_changes: Optional summary from auto_populate_employees_from_members

    Returns:
        Preview data with statistics and sample entries
    """
    meta = schedule_payload.get("meta", {})
    past_entries = schedule_payload.get("past_entries", [])
    future_entries = schedule_payload.get("future_entries", [])
    members = schedule_payload.get("members", [])

    # Count entry types in future entries
    shift_count = sum(1 for e in future_entries if e.get("entry_type") == "shift")
    timeoff_count = sum(1 for e in future_entries if e.get("entry_type") == "time_off")

    # Get unique employees in schedule
    schedule_employees = set()
    for entry in past_entries + future_entries:
        emp_name = entry.get("employee")
        if emp_name:
            schedule_employees.add(emp_name)

    # Sample entries for preview (first 10 future shifts)
    sample_shifts = [e for e in future_entries if e.get("entry_type") == "shift"][:10]
    sample_timeoff = [e for e in future_entries if e.get("entry_type") == "time_off"][:10]

    preview = {
        "summary": {
            "total_past_entries": len(past_entries),
            "total_future_entries": len(future_entries),
            "future_shifts": shift_count,
            "future_timeoff": timeoff_count,
            "members_in_file": len(members),
            "unique_employees_in_schedule": len(schedule_employees),
        },
        "metadata": meta,
        "sample_future_shifts": sample_shifts,
        "sample_future_timeoff": sample_timeoff,
    }

    # Add employee changes if provided
    if employee_changes:
        preview["employee_changes"] = {
            "added": employee_changes.get("added_count", 0),
            "already_existed": employee_changes.get("existing_count", 0),
            "total_processed": employee_changes.get("total_members", 0),
            "added_list": employee_changes.get("added_employees", []),
            "existing_list": employee_changes.get("existing_employees", []),
        }

    return preview

# ----------------------------
# Shift Pattern Detection
# ----------------------------
def detect_shift_patterns_from_schedule(
    schedule_payload: Dict[str, Any],
    project: Project
) -> Dict[str, Any]:
    """
    Analyze schedule data to automatically detect shift patterns and create shift templates.

    Args:
        schedule_payload: Parsed schedule data with past/future entries
        project: Current project (will be modified in place to add detected shifts)

    Returns:
        Summary dict with detected shift patterns and added shift templates
    """
    from models import ShiftTemplate
    from collections import defaultdict

    # Combine all entries for pattern analysis
    all_entries = schedule_payload.get("past_entries", []) + schedule_payload.get("future_entries", [])

    # Filter only shift entries (not time-off)
    shift_entries = [e for e in all_entries if e.get("entry_type") == "shift"]

    if not shift_entries:
        return {
            "detected_count": 0,
            "added_count": 0,
            "patterns": [],
            "message": "No shift entries found to analyze"
        }

    # Pattern detection based on: start_time + end_time + notes/label + color_code
    # Create a unique key for each pattern
    pattern_map = defaultdict(lambda: {
        "count": 0,
        "employees": set(),
        "dates": set(),
        "weekdays": set(),
        "color_code": None,
        "label": None,
        "notes": None,
        "unpaid_break": None,
        "example_entry": None
    })

    for entry in shift_entries:
        start_time = entry.get("start_time", "").strip()
        end_time = entry.get("end_time", "").strip()
        notes = (entry.get("notes") or "").strip()
        label = (entry.get("label") or "").strip()
        color_code = entry.get("color_code")

        # Skip entries without time information
        if not start_time or not end_time:
            continue

        # Create pattern key: start_time + end_time + notes (or label)
        # Use notes primarily, fallback to label
        # Use || as separator to avoid conflict with time format (HH:MM)
        role_identifier = notes or label or "Shift"
        pattern_key = f"{start_time}-{end_time}||{role_identifier}||{color_code or '1'}"

        pattern = pattern_map[pattern_key]
        pattern["count"] += 1

        # Track employee and date info
        emp_name = entry.get("employee")
        if emp_name:
            pattern["employees"].add(emp_name)

        entry_date = entry.get("start_date")
        if entry_date:
            pattern["dates"].add(entry_date)
            # Parse weekday
            try:
                dt = pd.to_datetime(entry_date)
                weekday_map = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}
                pattern["weekdays"].add(weekday_map[dt.weekday()])
            except:
                pass

        # Store pattern attributes
        if pattern["color_code"] is None:
            pattern["color_code"] = color_code
        if pattern["label"] is None and label:
            pattern["label"] = label
        if pattern["notes"] is None and notes:
            pattern["notes"] = notes
        if pattern["unpaid_break"] is None and entry.get("unpaid_break"):
            pattern["unpaid_break"] = entry.get("unpaid_break")

        # Store example entry for reference
        if pattern["example_entry"] is None:
            pattern["example_entry"] = entry

    # Convert patterns to shift templates
    # Only create patterns that occur more than once (to filter out one-offs)
    detected_patterns = []
    added_shifts = []

    for pattern_key, pattern in pattern_map.items():
        # Skip rare patterns (less than 2 occurrences)
        if pattern["count"] < 2:
            continue

        # Parse pattern key (format: "HH:MM-HH:MM||role||color")
        parts = pattern_key.split("||")
        if len(parts) < 3:
            continue

        time_part = parts[0]  # "HH:MM-HH:MM"
        role_part = parts[1]   # role/notes
        # parts[2] is color code (already stored in pattern dict)

        # Safely parse time_part - skip if malformed
        if "-" not in time_part:
            continue

        time_parts = time_part.split("-")
        if len(time_parts) != 2:
            continue

        start_time, end_time = time_parts[0].strip(), time_parts[1].strip()

        # Create shift ID from role and time
        shift_id = f"{role_part.lower().replace(' ', '-')}-{start_time.replace(':', '')}"

        # Check if shift already exists in project
        existing_shift = next((s for s in project.shifts if s.id == shift_id), None)

        if existing_shift:
            # Skip - already exists
            detected_patterns.append({
                "shift_id": shift_id,
                "role": role_part,
                "start_time": start_time,
                "end_time": end_time,
                "color_code": pattern["color_code"],
                "occurrences": pattern["count"],
                "weekdays": sorted(pattern["weekdays"]),
                "status": "already_exists"
            })
        else:
            # Create new shift template
            new_shift = ShiftTemplate(
                id=shift_id,
                role=role_part,
                start_time=start_time,
                end_time=end_time,
                weekdays=sorted(pattern["weekdays"]) if pattern["weekdays"] else ["Mon", "Tue", "Wed", "Thu", "Fri"],
                required_count={},  # Will be filled manually by user
                notes=pattern["notes"],
                color_code=pattern["color_code"],
                unpaid_break_minutes=pattern["unpaid_break"],
                teams_label=pattern["label"],
            )

            project.shifts.append(new_shift)
            added_shifts.append(new_shift)

            detected_patterns.append({
                "shift_id": shift_id,
                "role": role_part,
                "start_time": start_time,
                "end_time": end_time,
                "color_code": pattern["color_code"],
                "occurrences": pattern["count"],
                "weekdays": sorted(pattern["weekdays"]),
                "status": "added"
            })

    return {
        "detected_count": len(detected_patterns),
        "added_count": len(added_shifts),
        "patterns": detected_patterns,
        "message": f"Detected {len(detected_patterns)} shift patterns, added {len(added_shifts)} new shift templates"
    }

# ==============================================================================
# SCHEDULE ENTRY EXCEL EXPORT (Teams Format)
# ==============================================================================

def export_schedule_to_teams_excel(
    entries: List[Any],  # List[GeneratedScheduleEntry]
    members: List[Dict[str, str]],
    output_path: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> str:
    """
    Export schedule entries to Teams-compatible Excel format
    
    Args:
        entries: List of GeneratedScheduleEntry objects
        members: List of {name, email} dicts for Mitglieder sheet
        output_path: Path to save Excel file
        start_date: Optional filter start date
        end_date: Optional filter end date
    
    Returns:
        Path to created file
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Filter entries by date range if specified
    filtered_entries = entries
    if start_date or end_date:
        filtered_entries = []
        for entry in entries:
            try:
                entry_date = _parse_date(entry.start_date)
                if start_date and entry_date < start_date:
                    continue
                if end_date and entry_date > end_date:
                    continue
                filtered_entries.append(entry)
            except:
                filtered_entries.append(entry)  # Include if can't parse
    
    # Separate shifts and time-off
    shifts = [e for e in filtered_entries if e.entry_type == "shift"]
    time_offs = [e for e in filtered_entries if e.entry_type == "time_off"]
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Schichten (Shifts) sheet
    ws_shifts = wb.create_sheet("Schichten")
    _write_shifts_sheet(ws_shifts, shifts)
    
    # Create Arbeitsfreie Zeit (Time-Off) sheet
    ws_timeoff = wb.create_sheet("Arbeitsfreie Zeit")
    _write_timeoff_sheet(ws_timeoff, time_offs)
    
    # Create Mitglieder (Members) sheet
    ws_members = wb.create_sheet("Mitglieder")
    _write_members_sheet(ws_members, members)
    
    # Save workbook
    wb.save(output_path)
    return output_path


def _write_shifts_sheet(ws, shifts: List[Any]) -> None:
    """Write Schichten sheet with proper formatting"""
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Headers
    headers = [
        "Mitglied", "E-Mail (geschäftlich)", "Gruppe", "Startdatum", "Startzeit",
        "Enddatum", "Endzeit", "Themenfarbe", "Bezeichnung",
        "Unbezahlte Pause (Minuten)", "Notizen", "Geteilt"
    ]
    ws.append(headers)
    
    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Add data validation for color codes
    color_values = [
        "1. Weiß", "2. Blau", "3. Grün", "4. Lila", "5. Rosa", "6. Gelb",
        "8. Dunkelblau", "9. Dunkelgrün", "10. Dunkelviolett", "11. Dunkelrosa",
        "12. Dunkelgelb", "13. Grau"
    ]
    color_dv = DataValidation(type="list", formula1=f'"{",".join(color_values)}"', allow_blank=False)
    ws.add_data_validation(color_dv)
    
    # Add data validation for shared status
    shared_values = ["1. Geteilt", "2. Nicht freigegeben"]
    shared_dv = DataValidation(type="list", formula1=f'"{",".join(shared_values)}"', allow_blank=False)
    ws.add_data_validation(shared_dv)
    
    # Write shift data
    for entry in shifts:
        row = [
            entry.employee_name,
            entry.employee_email or "",
            entry.group or "",
            entry.start_date,
            entry.start_time,
            entry.end_date,
            entry.end_time,
            entry.color_code,
            entry.label or "",
            entry.unpaid_break,
            entry.notes or "",
            entry.shared
        ]
        ws.append(row)
        
        # Apply validation to this row
        row_num = ws.max_row
        color_dv.add(f"H{row_num}")
        shared_dv.add(f"L{row_num}")


def _write_timeoff_sheet(ws, time_offs: List[Any]) -> None:
    """Write Arbeitsfreie Zeit sheet with proper formatting"""
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Headers
    headers = [
        "Mitglied", "E-Mail (geschäftlich)", "Startdatum", "Startzeit",
        "Enddatum", "Endzeit", "Grund für arbeitsfreie Zeit", "Themenfarbe",
        "Notizen", "Geteilt"
    ]
    ws.append(headers)
    
    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Add data validation
    color_values = [
        "1. Weiß", "2. Blau", "3. Grün", "4. Lila", "5. Rosa", "6. Gelb",
        "8. Dunkelblau", "9. Dunkelgrün", "10. Dunkelviolett", "11. Dunkelrosa",
        "12. Dunkelgelb", "13. Grau"
    ]
    color_dv = DataValidation(type="list", formula1=f'"{",".join(color_values)}"', allow_blank=False)
    ws.add_data_validation(color_dv)
    
    shared_values = ["1. Geteilt", "2. Nicht freigegeben"]
    shared_dv = DataValidation(type="list", formula1=f'"{",".join(shared_values)}"', allow_blank=False)
    ws.add_data_validation(shared_dv)
    
    # Write time-off data
    for entry in time_offs:
        row = [
            entry.employee_name,
            entry.employee_email or "",
            entry.start_date,
            entry.start_time or "00:00",
            entry.end_date,
            entry.end_time or "00:00",
            entry.reason or "",
            entry.color_code if hasattr(entry, 'color_code') and entry.color_code else "",
            entry.notes or "",
            entry.shared
        ]
        ws.append(row)
        
        # Apply validation
        row_num = ws.max_row
        if entry.color_code:
            color_dv.add(f"H{row_num}")
        shared_dv.add(f"J{row_num}")


def _write_members_sheet(ws, members: List[Dict[str, str]]) -> None:
    """Write Mitglieder sheet"""
    from openpyxl.styles import Font
    
    # Headers
    ws.append(["Mitglied", "E-Mail (geschäftlich)"])
    
    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Write member data
    for member in members:
        ws.append([member.get("name", ""), member.get("email", "")])


def convert_uploaded_entries_to_schedule_entries(
    schedule_payload: Dict[str, Any],
    members: List[Dict[str, str]] = None
) -> Tuple[List[Any], List[Dict[str, str]]]:
    """
    Convert uploaded schedule payload to GeneratedScheduleEntry objects
    
    Args:
        schedule_payload: Parsed schedule from parse_teams_excel_multisheet
        members: Optional list of member dicts
    
    Returns:
        Tuple of (entries, members)
    """
    from models import GeneratedScheduleEntry
    
    entries = []

    # Process past entries (shifts and time-off)
    past_entries = schedule_payload.get("past_entries", schedule_payload.get("past", []))
    for past_entry in past_entries:
        entry = _payload_entry_to_schedule_entry(past_entry, source="uploaded")
        if entry:
            entries.append(entry)

    # Process future entries
    future_entries = schedule_payload.get("future_entries", schedule_payload.get("future", []))
    for future_entry in future_entries:
        entry = _payload_entry_to_schedule_entry(future_entry, source="uploaded")
        if entry:
            entries.append(entry)
    
    # Extract members if not provided
    if not members:
        members = schedule_payload.get("members", [])
    
    return entries, members


def _payload_entry_to_schedule_entry(payload_entry: Dict[str, Any], source: str = "uploaded") -> Optional[Any]:
    """Convert a payload entry dict to GeneratedScheduleEntry"""
    from models import GeneratedScheduleEntry
    
    try:
        # Determine entry type
        entry_type = payload_entry.get("entry_type", "shift")
        
        # Build entry data
        entry_data = {
            "employee_name": payload_entry.get("employee_name", ""),
            "employee_email": payload_entry.get("employee_email"),
            "group": payload_entry.get("group"),
            "start_date": payload_entry.get("start_date", ""),
            "start_time": payload_entry.get("start_time", "00:00"),
            "end_date": payload_entry.get("end_date", payload_entry.get("start_date", "")),
            "end_time": payload_entry.get("end_time", "00:00"),
            "color_code": payload_entry.get("color_code", "1. Weiß"),
            "label": payload_entry.get("label"),
            "unpaid_break": payload_entry.get("unpaid_break"),
            "notes": payload_entry.get("notes"),
            "shared": payload_entry.get("shared", "1. Geteilt"),
            "entry_type": entry_type,
            "reason": payload_entry.get("reason"),
            "source": source
        }
        
        return GeneratedScheduleEntry(**entry_data)
    except Exception as e:
        print(f"Error converting payload entry: {e}")
        return None

